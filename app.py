import io
import re
from copy import copy

import numpy as np
import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

# ============================================================
# PAGE
# ============================================================
st.set_page_config(page_title="Sales & Inventory Analysis", page_icon="📊", layout="wide")
st.title("📊 Sales & Inventory Analysis")
st.caption("Upload Sales, Inventory and Output Pattern files. The system automatically detects 2–6 months.")

MONTH_MAP = {
    "jan": 1, "january": 1, "feb": 2, "february": 2,
    "mar": 3, "march": 3, "apr": 4, "april": 4, "may": 5,
    "jun": 6, "june": 6, "jul": 7, "july": 7, "aug": 8,
    "august": 8, "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10, "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}


def norm(x):
    x = re.sub(r"[^a-z0-9]+", "_", str(x).strip().lower())
    return re.sub(r"_+", "_", x).strip("_")


def clean_text_series(s):
    return s.fillna("").astype(str).str.strip()


def standardize_subbrand_name(value):
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if not text:
        return ""
    return text.upper()


def find_col(df, names):
    cols = {norm(c): c for c in df.columns}
    wanted = [norm(n) for n in names]
    for n in wanted:
        if n in cols:
            return cols[n]
    # only use substring matching after exact matching
    for actual_norm, original in cols.items():
        for n in wanted:
            if n and (n in actual_norm or actual_norm in n):
                return original
    return None


def month_from_text(text):
    text = str(text).lower()
    for name, number in MONTH_MAP.items():
        if re.search(rf"\b{name}\b", text):
            m = re.search(r"(20\d{2})", text)
            year = int(m.group(1)) if m else 2026
            return pd.Timestamp(year=year, month=number, day=1)
    return None


def add_month(df, sheet_name="", file_name=""):
    df = df.copy()
    date_col = find_col(df, [
        "date", "sale_date", "sales_date", "transaction_date",
        "invoice_date", "bill_date", "order_date"
    ])
    if date_col:
        d = pd.to_datetime(df[date_col], errors="coerce")
        if d.notna().any():
            df["_month"] = d.dt.to_period("M").dt.to_timestamp()
            return df

    month_col = find_col(df, ["month", "sales_month", "sale_month", "transaction_month", "period"])
    if month_col:
        parsed = [month_from_text(x) for x in df[month_col]]
        if any(x is not None for x in parsed):
            df["_month"] = parsed
            return df

    detected = month_from_text(sheet_name) or month_from_text(file_name)
    if detected is not None:
        df["_month"] = detected
        return df
    return None


def clean_df(df):
    df = df.copy().dropna(axis=0, how="all").dropna(axis=1, how="all")
    df.columns = [str(c).strip() for c in df.columns]
    for c in df.columns:
        if df[c].dtype == "object":
            df[c] = df[c].apply(lambda x: x.strip() if isinstance(x, str) else x)
    return df


@st.cache_data(show_spinner=False)
def read_sales(file_bytes, file_name):
    frames = []
    stream = io.BytesIO(file_bytes)
    if file_name.lower().endswith(".csv"):
        df = clean_df(pd.read_csv(stream, low_memory=False))
        x = add_month(df, file_name=file_name)
        if x is not None:
            frames.append(x)
    else:
        book = pd.ExcelFile(stream)
        for sheet in book.sheet_names:
            df = clean_df(pd.read_excel(book, sheet_name=sheet))
            if df.empty:
                continue
            x = add_month(df, sheet_name=sheet, file_name=file_name)
            if x is not None:
                frames.append(x)
    if not frames:
        raise ValueError("Could not detect any month in the Sales file. Use a Date/Month column or month-named sheets.")
    out = pd.concat(frames, ignore_index=True, sort=False)
    return out[out["_month"].notna()].copy()


@st.cache_data(show_spinner=False)
def read_inventory(file_bytes, file_name):
    stream = io.BytesIO(file_bytes)
    if file_name.lower().endswith(".csv"):
        return clean_df(pd.read_csv(stream, low_memory=False))
    return clean_df(pd.read_excel(stream))


def prepare_sales(df):
    pcode = find_col(df, ["pcode", "product_code", "product_id", "sku"])
    qty = find_col(df, ["quantity", "qty", "sales_quantity", "sale_quantity"])
    value = find_col(df, ["net_amount", "sales_value", "sales_amount", "revenue", "amount"])
    if not pcode or not qty or not value:
        raise ValueError("Sales file must contain pcode, quantity and net_amount/sales value columns.")

    mapping = {pcode: "pcode", qty: "sales_quantity", value: "sales_value"}
    for source, target in [
        (find_col(df, ["prod_name", "product_name", "product"]), "prod_name"),
        (find_col(df, ["branch_name", "branch", "branch_location"]), "branch_name"),
        (find_col(df, ["branch_code", "branch_id"]), "branch_code"),
        (find_col(df, ["category", "product_category"]), "category"),
        (find_col(df, ["subbrandform_name", "subbrand", "sub_brand"]), "subbrandform_name"),
    ]:
        if source:
            mapping[source] = target

    sales = df.rename(columns=mapping).copy()
    sales["pcode"] = clean_text_series(sales["pcode"])
    sales["sales_quantity"] = pd.to_numeric(sales["sales_quantity"], errors="coerce").fillna(0)
    sales["sales_value"] = pd.to_numeric(sales["sales_value"], errors="coerce").fillna(0)
    if "subbrandform_name" in sales.columns:
        sales["subbrandform_name"] = sales["subbrandform_name"].map(standardize_subbrand_name)
    sales["branch_name"] = clean_text_series(sales["branch_name"]) if "branch_name" in sales.columns else "All Branches"
    if "branch_code" in sales.columns:
        sales["branch_code"] = clean_text_series(sales["branch_code"])
    return sales


def prepare_inventory(df):
    pcode = find_col(df, ["pcode", "product_code", "product_id", "sku"])
    qty = find_col(df, ["quantity", "stock", "inventory", "inventory_quantity"])
    if not pcode or not qty:
        raise ValueError("Inventory file must contain pcode and Quantity/Inventory columns.")

    mapping = {pcode: "pcode", qty: "inventory_quantity"}
    for source, target in [
        (find_col(df, ["branch_name", "branch_location", "branch"]), "inventory_branch_name"),
        (find_col(df, ["branch_code", "branch_id"]), "inventory_branch_code"),
        (find_col(df, ["prod_name", "product_name", "product"]), "inventory_prod_name"),
        (find_col(df, ["category", "product_category"]), "inventory_category"),
        (find_col(df, ["subbrandform_name", "subbrand", "sub_brand"]), "inventory_subbrandform_name"),
    ]:
        if source:
            mapping[source] = target

    inv = df.rename(columns=mapping).copy()
    inv["pcode"] = clean_text_series(inv["pcode"])
    inv["inventory_quantity"] = pd.to_numeric(inv["inventory_quantity"], errors="coerce").fillna(0)
    if "inventory_subbrandform_name" in inv.columns:
        inv["inventory_subbrandform_name"] = inv["inventory_subbrandform_name"].map(standardize_subbrand_name)
    if "inventory_branch_name" in inv.columns:
        inv["inventory_branch_name"] = clean_text_series(inv["inventory_branch_name"])
    if "inventory_branch_code" in inv.columns:
        inv["inventory_branch_code"] = clean_text_series(inv["inventory_branch_code"])
    return inv


def calculate_risk(inventory_qty, average_qty, monthly_values, month_labels):
    """Rules verified from the supplied original output."""
    inventory_qty = float(inventory_qty)
    average_qty = float(average_qty)

    # Zero/negative average sales are High Risk, even when inventory is zero.
    if average_qty <= 0:
        ratio = np.nan if inventory_qty == 0 else average_qty / inventory_qty
        return "High Risk", "Zero sale", ratio

    # Positive sales with zero inventory: original output shows Low Risk and blank Sales %.
    if inventory_qty == 0:
        return "Low Risk", "", np.nan

    ratio = average_qty / inventory_qty

    if ratio < 0.20:
        values = [float(v) if pd.notna(v) else 0 for v in monthly_values]
        idx = int(np.argmax(values))
        highest_qty = values[idx]
        qty_text = f"{int(highest_qty):,}" if float(highest_qty).is_integer() else f"{highest_qty:,.2f}"
        remark = f"Below 20% sale; highest sale in {month_labels[idx]} ({qty_text})"
        return "High Risk", remark, ratio
    if ratio < 0.50:
        return "Medium Risk", "", ratio
    return "Low Risk", "", ratio


def first_nonblank(series):
    for x in series:
        if pd.notna(x) and str(x).strip():
            return x
    return ""


def build_product_report(sales, inventory, months, month_labels):
    # Inventory is the base population: exactly one row per inventory pcode.
    inv_base = inventory.groupby("pcode", as_index=False, sort=False)["inventory_quantity"].sum()

    inv_meta = inventory.groupby("pcode", as_index=False, sort=False).agg(
        inventory_prod_name=("inventory_prod_name", first_nonblank) if "inventory_prod_name" in inventory.columns else ("pcode", lambda x: ""),
        inventory_category=("inventory_category", first_nonblank) if "inventory_category" in inventory.columns else ("pcode", lambda x: ""),
        inventory_subbrandform_name=("inventory_subbrandform_name", first_nonblank) if "inventory_subbrandform_name" in inventory.columns else ("pcode", lambda x: ""),
    )

    sales_meta_cols = ["pcode"] + [c for c in ["prod_name", "category", "subbrandform_name"] if c in sales.columns]
    sales_meta = sales[sales_meta_cols].groupby("pcode", as_index=False, sort=False).agg({c: first_nonblank for c in sales_meta_cols if c != "pcode"})

    report = inv_base.merge(inv_meta, on="pcode", how="left")
    report = report.merge(sales_meta, on="pcode", how="left")

    report["prod_name"] = report.get("inventory_prod_name", "").replace("", np.nan).fillna(report.get("prod_name", ""))
    report["category"] = report.get("inventory_category", "").replace("", np.nan).fillna(report.get("category", ""))
    report["subbrandform_name"] = report.get("inventory_subbrandform_name", "").replace("", np.nan).fillna(report.get("subbrandform_name", ""))
    if "subbrandform_name" in report.columns:
        report["subbrandform_name"] = report["subbrandform_name"].map(standardize_subbrand_name)

    monthly = pd.pivot_table(sales, index="pcode", columns="_month", values="sales_quantity", aggfunc="sum", fill_value=0)
    for month, label in zip(months, month_labels):
        report[label] = report["pcode"].map(monthly[month]).fillna(0) if month in monthly.columns else 0

    n = len(month_labels)
    report["Past sale Qty"] = report[month_labels].sum(axis=1)
    report["Average sale Qty"] = report["Past sale Qty"] / n
    values = sales[sales["_month"].isin(months)].groupby("pcode")["sales_value"].sum()
    report["Past sale Value"] = report["pcode"].map(values).fillna(0)
    report["Average sale Value"] = report["Past sale Value"] / n

    risks, remarks, ratios = [], [], []
    for _, row in report.iterrows():
        risk, remark, ratio = calculate_risk(row["inventory_quantity"], row["Average sale Qty"], [row[m] for m in month_labels], month_labels)
        risks.append(risk); remarks.append(remark); ratios.append(ratio)

    report["Risk assesment"] = risks
    report["Remarks"] = remarks
    report["Sales %"] = ratios
    report["Inventory Total Quantity"] = report["inventory_quantity"]

    qlabel = f"Past {n} months sale Qty"
    vlabel = f"Past {n} months sale Value"
    report[qlabel] = report["Past sale Qty"]
    report[vlabel] = report["Past sale Value"]

    columns = ["pcode", "prod_name", "category", "subbrandform_name", "Inventory Total Quantity", "Sales %", "Risk assesment", "Remarks"] + month_labels + [qlabel, vlabel, "Average sale Qty", "Average sale Value"]
    return report[columns].copy(), qlabel, vlabel


def build_branch_report(sales, inventory, months, month_labels):
    # Aggregate SALES by pcode + branch. This is only for the dashboard when branch is selected.
    keys = ["pcode", "branch_name"]
    if "branch_code" in sales.columns:
        keys.append("branch_code")

    branch = sales.groupby(keys, as_index=False).agg(
        total_quantity=("sales_quantity", "sum"),
        total_value=("sales_value", "sum")
    )

    monthly = pd.pivot_table(sales, index=keys, columns="_month", values="sales_quantity", aggfunc="sum", fill_value=0).reset_index()
    branch = branch.merge(monthly, on=keys, how="left")

    # Product metadata comes from sales; inventory metadata is used as fallback.
    meta = sales.groupby("pcode", as_index=False).agg({c: first_nonblank for c in ["prod_name", "category", "subbrandform_name"] if c in sales.columns})
    branch = branch.merge(meta, on="pcode", how="left")

    for month, label in zip(months, month_labels):
        branch[label] = branch[month] if month in branch.columns else 0
        if month in branch.columns:
            branch.drop(columns=[month], inplace=True)

    # Attach branch inventory. Prefer branch_code matching; otherwise normalized branch name.
    if "inventory_branch_code" in inventory.columns and "branch_code" in branch.columns:
        inv_b = inventory.groupby(["pcode", "inventory_branch_code"], as_index=False)["inventory_quantity"].sum()
        branch = branch.merge(inv_b, left_on=["pcode", "branch_code"], right_on=["pcode", "inventory_branch_code"], how="left")
        branch["Inventory Total Quantity"] = branch["inventory_quantity"]
        branch.drop(columns=["inventory_branch_code", "inventory_quantity"], inplace=True, errors="ignore")
    elif "inventory_branch_name" in inventory.columns:
        inv_b = inventory.copy()
        inv_b["_branch_key"] = inv_b["inventory_branch_name"].map(norm)
        branch["_branch_key"] = branch["branch_name"].map(norm)
        inv_b = inv_b.groupby(["pcode", "_branch_key"], as_index=False)["inventory_quantity"].sum()
        branch = branch.merge(inv_b, on=["pcode", "_branch_key"], how="left")
        branch["Inventory Total Quantity"] = branch["inventory_quantity"]
        branch.drop(columns=["_branch_key", "inventory_quantity"], inplace=True, errors="ignore")
    else:
        branch["Inventory Total Quantity"] = np.nan

    # Dashboard risk is calculated from the branch's own sales and branch inventory.
    avg = branch["total_quantity"] / len(month_labels)
    risks, remarks, ratios = [], [], []
    for i, row in branch.iterrows():
        risk, remark, ratio = calculate_risk(row["Inventory Total Quantity"] if pd.notna(row["Inventory Total Quantity"]) else 0, avg.loc[i], [row[m] for m in month_labels], month_labels)
        risks.append(risk); remarks.append(remark); ratios.append(ratio)
    branch["Risk assesment"] = risks
    branch["Remarks"] = remarks
    branch["Sales %"] = ratios
    return branch


def build_reports(sales, inventory):
    months = sorted(pd.Timestamp(x) for x in sales["_month"].dropna().unique())
    if len(months) > 6:
        months = months[-6:]
    if not 2 <= len(months) <= 6:
        raise ValueError(f"{len(months)} months detected. Please provide between 2 and 6 months.")
    month_labels = [m.strftime("%b'%y").replace("Jul'", "July'") for m in months]
    report, qlabel, vlabel = build_product_report(sales, inventory, months, month_labels)
    branch_report = build_branch_report(sales, inventory, months, month_labels)
    return report, branch_report, months, month_labels, qlabel, vlabel


def generate_excel(report, pattern_bytes, month_labels):
    wb = load_workbook(io.BytesIO(pattern_bytes))
    if wb.sheetnames and wb.sheetnames[0] != "Inventory Analysis":
        wb.active.title = "Inventory Analysis"
    if "Inventory Analysis" not in wb.sheetnames:
        wb.active.title = "Inventory Analysis"
    if "Summary" not in wb.sheetnames:
        wb.create_sheet("Summary")

    ws = wb["Inventory Analysis"]

    # Remove existing merges safely before rewriting the dynamic month group.
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))

    n = len(month_labels)
    qlabel = f"Past {n} months sale Qty"
    vlabel = f"Past {n} months sale Value"
    headers = ["pcode", "prod_name", "category", "subbrandform_name", "Inventory Total Quantity", "Sales %", "Risk assesment", "Remarks"] + month_labels + [vlabel, "Average sale Qty", "Average sale Value"]
    required = len(headers)

    # Preserve the pattern's first two rows; replace data rows only.
    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)
    if ws.max_column > required:
        ws.delete_cols(required + 1, ws.max_column - required)
    elif ws.max_column < required:
        ws.insert_cols(ws.max_column + 1, required - ws.max_column)

    for c, header in enumerate(headers, 1):
        ws.cell(2, c).value = header

    # Dynamic month heading in row 1.
    ws.merge_cells(start_row=1, start_column=9, end_row=1, end_column=8 + n)
    ws.cell(1, 9).value = qlabel

    for r, (_, row) in enumerate(report.iterrows(), 3):
        for c, header in enumerate(headers, 1):
            value = row.get(header, "")
            if pd.isna(value):
                value = ""
            cell = ws.cell(r, c)
            cell.value = value
            if header == "Sales %":
                cell.number_format = "0.000000"
            elif header in month_labels or header in ["Inventory Total Quantity", "Average sale Qty"]:
                cell.number_format = "#,##0.00"
            elif header in [vlabel, "Average sale Value"]:
                cell.number_format = "#,##0.00"

    ws.freeze_panes = "A3"
    for col_num in range(1, required + 1):
        max_len = 0
        for row_num in range(1, ws.max_row + 1):
            cell = ws.cell(row_num, col_num)
            if isinstance(cell, MergedCell):
                continue
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_num)].width = min(max(max_len + 2, 10), 35)

    summary_ws = wb["Summary"]
    summary_ws.delete_rows(1, summary_ws.max_row)
    summary_ws["A1"] = "Metric"
    summary_ws["B1"] = "Value"
    mapped_sales = set(sales["pcode"].dropna().astype(str).str.strip())
    summary_ws["A2"] = "Unique inventory PCodes"
    summary_ws["B2"] = report["pcode"].nunique()
    summary_ws["A3"] = "Total rows in analysis"
    summary_ws["B3"] = len(report)
    summary_ws["A4"] = "PCodes with mapped sales"
    summary_ws["B4"] = int(report["pcode"].isin(mapped_sales).sum())
    summary_ws["A5"] = "PCodes with zero/negative average sale"
    summary_ws["B5"] = int((report.get("Average sale Qty", 0) <= 0).sum())
    summary_ws["A6"] = "PCodes below 20% Sales %"
    summary_ws["B6"] = int((report.get("Sales %", 0) < 0.20).sum())
    summary_ws["A7"] = "High Risk"
    summary_ws["B7"] = int((report.get("Risk assesment", "") == "High Risk").sum())
    summary_ws["A8"] = "Medium Risk"
    summary_ws["B8"] = int((report.get("Risk assesment", "") == "Medium Risk").sum())
    summary_ws["A9"] = "Low Risk"
    summary_ws["B9"] = int((report.get("Risk assesment", "") == "Low Risk").sum())
    summary_ws.freeze_panes = "A2"
    for col_num in range(1, 3):
        summary_ws.column_dimensions[get_column_letter(col_num)].width = 28

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ============================================================
# UPLOADS
# ============================================================
st.sidebar.header("📂 Upload Files")
sales_file = st.sidebar.file_uploader("1️⃣ Sales File", type=["xlsx", "xls", "xlsm", "csv"])
inventory_file = st.sidebar.file_uploader("2️⃣ Inventory File", type=["xlsx", "xls", "xlsm", "csv"])
pattern_file = st.sidebar.file_uploader("3️⃣ Output Pattern File", type=["xlsx", "xlsm"])

if not (sales_file and inventory_file and pattern_file):
    st.info("Upload all three files to start the analysis.")
    st.stop()

try:
    with st.spinner("Processing Sales and Inventory..."):
        sales = prepare_sales(read_sales(sales_file.getvalue(), sales_file.name))
        inventory = prepare_inventory(read_inventory(inventory_file.getvalue(), inventory_file.name))
        report, branch_report, months, month_labels, qlabel, vlabel = build_reports(sales, inventory)
except Exception as e:
    st.error(f"❌ Processing failed: {e}")
    st.stop()

st.success("✅ " + f"{len(months)} months detected: " + ", ".join(pd.Timestamp(m).strftime("%B %Y") for m in months))

# ============================================================
# DASHBOARD
# ============================================================
st.header("📊 Dashboard")

# Start from PRODUCT-LEVEL data. Only switch to branch-level after a branch is selected.
f1, f2, f3, f4 = st.columns(4)

branches = sorted(branch_report["branch_name"].dropna().astype(str).unique())
selected_branches = f1.multiselect("🏢 Branch", branches)

if selected_branches:
    filtered = branch_report[branch_report["branch_name"].isin(selected_branches)].copy()
else:
    filtered = report.copy()

# Category options must come from the CURRENT base dataset, not an empty column.
categories = sorted(filtered["category"].replace("", np.nan).dropna().astype(str).unique()) if "category" in filtered.columns else []
selected_categories = f2.multiselect("📦 Category", categories)
if selected_categories:
    filtered = filtered[filtered["category"].isin(selected_categories)].copy()

risk_values = sorted(filtered["Risk assesment"].dropna().astype(str).unique())
selected_risks = f3.multiselect("⚠️ Risk", risk_values)
if selected_risks:
    filtered = filtered[filtered["Risk assesment"].isin(selected_risks)].copy()

feature = f4.selectbox("📈 Feature", ["Sales Value", "Total Quantity", "Inventory", "Sales %", "Risk Analysis"])

# Standardize dashboard metrics for both product and branch modes.
if selected_branches:
    quantity_col = "total_quantity"
    value_col = "total_value"
else:
    quantity_col = qlabel
    value_col = vlabel

k1, k2, k3, k4 = st.columns(4)
k1.metric("Products / Rows", f"{filtered['pcode'].nunique():,}")
k2.metric("Total Quantity", f"{filtered[quantity_col].sum():,.0f}")
k3.metric("Sales Value", f"{filtered[value_col].sum():,.2f}")
if "Inventory Total Quantity" in filtered.columns:
    inventory_series = pd.to_numeric(filtered["Inventory Total Quantity"], errors="coerce").fillna(0)
else:
    inventory_series = pd.Series(0, index=filtered.index)
inv_total = inventory_series.sum()
k4.metric("Inventory", f"{inv_total:,.0f}")

# ============================================================
# FEATURE CHART
# ============================================================
st.subheader("📈 Feature Chart")

if selected_branches:
    group_col = "branch_name"
    if feature == "Sales Value":
        chart = filtered.groupby(group_col, as_index=False)[value_col].sum().sort_values(value_col, ascending=False).head(20)
        fig = px.bar(chart, x=group_col, y=value_col, title="Sales Value by Branch")
    elif feature == "Total Quantity":
        chart = filtered.groupby(group_col, as_index=False)[quantity_col].sum().sort_values(quantity_col, ascending=False).head(20)
        fig = px.bar(chart, x=group_col, y=quantity_col, title="Total Quantity by Branch")
    elif feature == "Inventory":
        chart = filtered.groupby(group_col, as_index=False)["Inventory Total Quantity"].sum().sort_values("Inventory Total Quantity", ascending=False).head(20)
        fig = px.bar(chart, x=group_col, y="Inventory Total Quantity", title="Inventory by Branch")
    elif feature == "Sales %":
        chart = filtered.groupby(group_col, as_index=False)["Sales %"].mean().sort_values("Sales %", ascending=False).head(20)
        fig = px.bar(chart, x=group_col, y="Sales %", title="Average Sales % by Branch")
    else:
        chart = filtered["Risk assesment"].value_counts().rename_axis("Risk").reset_index(name="Products")
        fig = px.pie(chart, names="Risk", values="Products", title="Risk Analysis")
else:
    # No branch selected: chart the PRODUCT-LEVEL population.
    if feature == "Sales Value":
        chart = filtered.groupby("category", as_index=False)[value_col].sum().sort_values(value_col, ascending=False).head(20)
        fig = px.bar(chart, x="category", y=value_col, title="Sales Value by Category")
    elif feature == "Total Quantity":
        chart = filtered.groupby("category", as_index=False)[quantity_col].sum().sort_values(quantity_col, ascending=False).head(20)
        fig = px.bar(chart, x="category", y=quantity_col, title="Total Quantity by Category")
    elif feature == "Inventory":
        chart = filtered.groupby("category", as_index=False)["Inventory Total Quantity"].sum().sort_values("Inventory Total Quantity", ascending=False).head(20)
        fig = px.bar(chart, x="category", y="Inventory Total Quantity", title="Inventory by Category")
    elif feature == "Sales %":
        chart = filtered.groupby("category", as_index=False)["Sales %"].mean().sort_values("Sales %", ascending=False).head(20)
        fig = px.bar(chart, x="category", y="Sales %", title="Average Sales % by Category")
    else:
        chart = filtered["Risk assesment"].value_counts().rename_axis("Risk").reset_index(name="Products")
        fig = px.pie(chart, names="Risk", values="Products", title="Risk Analysis")

st.plotly_chart(fig, use_container_width=True)

# Monthly chart always uses the currently filtered dataset.
st.subheader("📅 Monthly Sales Quantity")
monthly_chart = pd.DataFrame({"Month": month_labels, "Quantity": [filtered[m].sum() for m in month_labels]})
fig_monthly = px.bar(monthly_chart, x="Month", y="Quantity", title=f"Sales Quantity – Past {len(month_labels)} Months")
st.plotly_chart(fig_monthly, use_container_width=True)

# ============================================================
# TABLE
# ============================================================
st.subheader("📋 Filtered Analysis")
if selected_branches:
    table_cols = ["branch_name", "pcode"] + (["category"] if "category" in filtered.columns else []) + month_labels + ["total_quantity", "total_value", "Inventory Total Quantity", "Sales %", "Risk assesment"]
else:
    table_cols = ["pcode", "prod_name", "category", "subbrandform_name", "Inventory Total Quantity", "Sales %", "Risk assesment", "Remarks"] + month_labels + ["Past sale Value", "Average sale Qty", "Average sale Value"]

table_cols = [c for c in table_cols if c in filtered.columns]
st.dataframe(filtered[table_cols], use_container_width=True, height=450)

# ============================================================
# FINAL EXCEL
# ============================================================
st.header("📥 Final Excel")
st.write("The final Excel contains one row per inventory pcode and follows the uploaded output pattern. Dashboard filters do not change the final Excel.")

if st.button("Generate Final Excel", type="primary"):
    try:
        with st.spinner("Creating final Excel..."):
            excel_bytes = generate_excel(report, pattern_file.getvalue(), month_labels)
        st.success(f"✅ Final Excel created with {len(report):,} inventory PCodes.")
        st.download_button("⬇️ Download Final Excel", data=excel_bytes, file_name="Sales_Inventory_Analysis.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        st.error(f"❌ Excel generation failed: {e}")

